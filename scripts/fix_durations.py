import re, json, subprocess
from openpyxl import load_workbook

HTML = r'C:\Users\aelha\Downloads\gimi-store.html'
BS=chr(92); SQ=chr(39)

# --- 1. Sheet Hours (source for 29 courses) ---
wb = load_workbook(r'C:\Users\aelha\Downloads\Copy of Erila_ GIMI_CRM_Database_v3.xlsx', data_only=True)
ws = wb['Schemes']; H=[c.value for c in ws[1]]
sheet_hours = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    r=dict(zip(H,row)); sheet_hours[str(r['Code']).strip()] = r.get('Hours')

def fmt_hours(v):
    if v is None: return 'TBD'
    s=str(v).strip()
    if s=='' or s.lower()=='none': return 'TBD'
    if 'month' in s.lower(): return s            # "6 months"
    if s.lower()=='variable': return 'Variable'
    if '-' in s: return f'{s}h'                  # "32-40" -> "32-40h"
    try:
        n=float(s); return f'{int(n)}h' if n==int(n) else f'{n}h'
    except: return s

# --- 2. Deck durations for masterclasses ---
DECK_DUR = {'CIP-MC':'8 weeks', 'CCIO-MC':'4 weeks'}

# --- 3. Build target duration per code ---
COURSE_CODES = ['IP','PBCC','CGIS-1','CIP-0','CIP-1','CIP-2','CCIO-1','CCIO-2','CGIT','CGT',
                'CDT-1','CDT-2','CDTT-3','CFF-1','CFF-2','CFF-3','CFF-4','CFF-5','CLF-1','CIL',
                'MCI-1','MCI-2','MCI-3','MCI-4','CAAP','CGA','CIME','CTC','CLC']
target = {}
for c in COURSE_CODES:
    target[c] = fmt_hours(sheet_hours.get(c))
target['CIP-MC'] = DECK_DUR['CIP-MC']
target['CCIO-MC'] = DECK_DUR['CCIO-MC']

CODE_TO_ID = {
 'IP':'ip','PBCC':'pbcc','CGIS-1':'cgis-1','CIP-0':'cip-0','CIP-1':'cip-1','CIP-2':'cip-2',
 'CIP-MC':'cip-mc','CCIO-1':'ccio-1','CCIO-2':'ccio-2','CCIO-MC':'ccio-mc','CGIT':'cgit','CGT':'cgt',
 'CDT-1':'cdt-1','CDT-2':'cdt-2','CDTT-3':'cdtt-3','CFF-1':'cff-1','CFF-2':'cff-2','CFF-3':'cff-3',
 'CFF-4':'cff-4','CFF-5':'cff-5','CLF-1':'clf-1','CIL':'cil','MCI-1':'mci-1','MCI-2':'mci-2',
 'MCI-3':'mci-3','MCI-4':'mci-4','CAAP':'caap','CGA':'cga','CIME':'cime','CTC':'ctc','CLC':'clc',
}

html = open(HTML, encoding='utf-8').read()

def get_block(pid):
    pat = re.compile(r"mk\(\{ id:'" + re.escape(pid) + r"',.*?\}\)", re.DOTALL)
    m = pat.search(html); return m.group(0) if m else None

def cur_duration(block):
    key="duration:"+SQ; pos=block.find(key)
    if pos<0: return None
    k=pos+len(key); out=[]
    while k<len(block):
        c=block[k]
        if c==SQ: break
        out.append(c); k+=1
    return ''.join(out)

changes=[]
for code in CODE_TO_ID:
    pid = CODE_TO_ID[code]
    block = get_block(pid)
    if not block: print('NO BLOCK', code); continue
    cur = cur_duration(block)
    tgt = target[code]
    if cur != tgt:
        new_block = re.sub(r"duration:'[^']*'", f"duration:'{tgt}'", block, count=1)
        html = html.replace(block, new_block)
        changes.append((code, cur, tgt))

open(HTML,'w',encoding='utf-8').write(html)
print(f'Duration changes applied: {len(changes)}')
for code,old,new in changes:
    print(f'  {code:8} {old!r:12} -> {new!r}')

r = subprocess.run(['node','-e',
    "const fs=require('fs');const h=fs.readFileSync(String.raw`"+HTML+"`,'utf8');"
    "const m=h.match(/<script>([\s\S]*?)<\/script>/);"
    "try{new Function(m[1]);console.log('JS OK')}catch(e){console.log('Err:',e.message)}"],
    capture_output=True, text=True)
print(r.stdout.strip())
