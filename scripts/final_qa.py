import re, json
from openpyxl import load_workbook
BS=chr(92); SQ=chr(39)
html=open('gimi-store.html',encoding='utf-8').read()

def parse_blocks(s):
    res=[]; i=0
    while True:
        idx=s.find('mk({', i)
        if idx<0: break
        depth=1; j=idx+3
        while j<len(s) and depth>0:
            j+=1
            if s[j]=='{':depth+=1
            elif s[j]=='}':depth-=1
        res.append(s[idx+3:j+1]); i=s.find(')',j)+1
    return res
def field(b,name):
    key=name+":"+SQ; pos=b.find(key)
    if pos<0: return None
    k=pos+len(key);out=[]
    while k<len(b):
        c=b[k]
        if c==BS: out.append(b[k+1]); k+=2; continue
        if c==SQ: break
        out.append(c); k+=1
    return ''.join(out)

ID2CODE={'ip':'IP','pbcc':'PBCC','cgis-1':'CGIS-1','cip-0':'CIP-0','cip-1':'CIP-1','cip-2':'CIP-2',
 'cip-mc':'CIP-MC','ccio-1':'CCIO-1','ccio-2':'CCIO-2','ccio-mc':'CCIO-MC','cgit':'CGIT','cgt':'CGT',
 'cdt-1':'CDT-1','cdt-2':'CDT-2','cdtt-3':'CDTT-3','cff-1':'CFF-1','cff-2':'CFF-2','cff-3':'CFF-3',
 'cff-4':'CFF-4','cff-5':'CFF-5','clf-1':'CLF-1','cil':'CIL','mci-1':'MCI-1','mci-2':'MCI-2',
 'mci-3':'MCI-3','mci-4':'MCI-4','caap':'CAAP','cga':'CGA','cime':'CIME','ctc':'CTC','clc':'CLC'}
store={}
for b in parse_blocks(re.search(r'const PRODUCTS = \[([\s\S]*?)\n\];',html).group(1)):
    pid=field(b,'id')
    if pid in ID2CODE:
        store[ID2CODE[pid]]={'cert':field(b,'certDisplay'),'train':field(b,'trainDisplay'),
                             'dur':field(b,'duration'),'desc':field(b,'description') or ''}

deck=json.load(open('deck_prices.json',encoding='utf-8'))
crm=json.load(open('crm_v3_2.json',encoding='utf-8'))
wb=load_workbook(r'C:\Users\aelha\Downloads\Copy of Erila_ GIMI_CRM_Database_v3.xlsx',data_only=True)
ws=wb['Schemes'];Hh=[c.value for c in ws[1]]
shours={}
for row in ws.iter_rows(min_row=2,values_only=True):
    if not row[0]:continue
    r=dict(zip(Hh,row)); shours[str(r['Code']).strip()]=r.get('Hours')
def fmt(v):
    if v is None:return 'TBD'
    s=str(v).strip()
    if s=='' or s.lower()=='none':return 'TBD'
    if 'month' in s.lower():return s
    if s.lower()=='variable':return 'Variable'
    if '-' in s:return f'{s}h'
    try:
        n=float(s);return f'{int(n)}h' if n==int(n) else f'{n}h'
    except:return s
def npr(s):return None if s is None else str(s).replace('*','').strip()
def ntx(s):return ' '.join((s or '').split())

DECK_DUR={'CIP-MC':'8 weeks','CCIO-MC':'4 weeks'}
DECK_DESC={'CIP-MC','CCIO-MC'}
order=['CIP-0','CIP-1','CIP-2','CIP-MC','CCIO-1','CCIO-2','CCIO-MC','CGIS-1','CGIT','CGT','IP','PBCC',
 'CDT-1','CDT-2','CDTT-3','CAAP','CGA','CIME','CTC','CLC','CFF-1','CFF-2','CFF-3','CFF-4','CFF-5',
 'CLF-1','CIL','MCI-1','MCI-2','MCI-3','MCI-4']
bad=0
for code in order:
    s=store[code]
    # price
    pc = '' if npr(s['cert'])==npr(deck[code].get('cert')) else 'CERTPRICE '
    pt = '' if npr(s['train'])==npr(deck[code].get('train')) else 'TRAINPRICE '
    # duration target
    dtgt = DECK_DUR[code] if code in DECK_DUR else fmt(shours.get(code))
    pd = '' if s['dur']==dtgt else f'DUR(store={s["dur"]} want={dtgt}) '
    # desc target
    if code in DECK_DESC: dsrc=ntx(deck[code].get('desc'))
    else: dsrc=ntx(crm.get(code,{}).get('description'))
    dd = '' if ntx(s['desc'])==dsrc else 'DESC '
    flags=(pc+pt+pd+dd).strip()
    if flags: bad+=1; print(f'  [{code:7}] {flags}')
print('\nRESULT:', 'ALL 31 PRODUCTS FULLY ALIGNED' if bad==0 else f'{bad} products still off')
